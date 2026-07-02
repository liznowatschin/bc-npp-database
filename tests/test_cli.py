import json

from openpyxl import Workbook
from typer.testing import CliRunner

from bc_npp_database.cli import app

runner = CliRunner()


def test_cli_version():
    result = runner.invoke(app, ["--version"])
    assert result.exit_code == 0
    assert "bc-npp-database 0.1.0a0" in result.stdout


def test_cli_info():
    result = runner.invoke(app, ["info"])
    assert result.exit_code == 0
    assert "BC Native Plant & Pollinator Database" in result.stdout
    assert "BC-NPPD" in result.stdout


def test_validate_source_policy_command_detects_excluded_url(tmp_path):
    source_file = tmp_path / "source.txt"
    source_file.write_text(
        "https://vancouver.ca/files/cov/vancouver-gri-planting-guidelines.pdf",
        encoding="utf-8",
    )

    result = runner.invoke(app, ["validate-source-policy", str(source_file)])

    assert result.exit_code == 1
    assert "Excluded source found" in result.stderr


def test_inventory_workbook_json_command(tmp_path):
    workbook_path = tmp_path / "fixture.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Species_Master"
    worksheet.append(["Species_ID", "Botanical_Name"])
    workbook.save(workbook_path)

    result = runner.invoke(app, ["inventory-workbook", str(workbook_path), "--json"])

    assert result.exit_code == 0
    assert '"Species_Master"' in result.stdout
    assert '"Species_ID"' in result.stdout


def test_validate_workbook_command_reports_missing_sheets(tmp_path):
    workbook_path = tmp_path / "fixture.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Species_Master"
    worksheet.append(["Species_ID"])
    workbook.save(workbook_path)

    result = runner.invoke(app, ["validate-workbook", str(workbook_path), "--json"])

    assert result.exit_code == 1
    assert "missing_sheet" in result.stdout


def test_validate_source_records_command_json(tmp_path):
    source_path = tmp_path / "sources.csv"
    source_path.write_text(
        "source_id,source_name,source_tier,citation\n"
        "SRC-0001,LMH77,Tier 2,Klassen et al. 2026.\n",
        encoding="utf-8",
    )

    result = runner.invoke(app, ["validate-source-records", str(source_path), "--json"])

    assert result.exit_code == 0
    assert result.stdout.strip() == "[]"


def test_validate_source_attribution_command_json_reports_errors(tmp_path):
    source_path = tmp_path / "attribution.json"
    source_path.write_text(
        '[{"source_id": "SRC-0001", "claim_field": "habitat", '
        '"evidence_confidence": "Certain"}]',
        encoding="utf-8",
    )

    result = runner.invoke(app, ["validate-source-attribution", str(source_path), "--json"])

    assert result.exit_code == 1
    assert "invalid_evidence_confidence" in result.stdout
    assert "missing_required_field" in result.stdout


def test_import_canonical_workbook_command_json(tmp_path):
    workbook_path = tmp_path / "canonical.xlsx"
    workbook = Workbook()
    species = workbook.active
    species.title = "Species_Master"
    species.append(["Species_ID", "Botanical_Name", "Evidence_Confidence"])
    species.append(["BCNPPD-0001", "Achillea millefolium", "A"])
    lookups = workbook.create_sheet("Lookup_Tables")
    lookups.append(["lookup_name", "value"])
    lookups.append(["Life Cycle", "Perennial"])
    attribution = workbook.create_sheet("Source_Attribution")
    attribution.append(["Source ID", "Species_ID", "Field", "Confidence"])
    attribution.append(["SRC-0001", "BCNPPD-0001", "habitat", "B"])
    bloom = workbook.create_sheet("Bloom_Calendar")
    bloom.append(["Species_ID", "Common_Name", "Jan"])
    bloom.append(["BCNPPD-0001", "yarrow", ""])
    workbook.save(workbook_path)

    result = runner.invoke(app, ["import-canonical-workbook", str(workbook_path), "--json"])

    assert result.exit_code == 0
    assert '"species": 1' in result.stdout
    assert '"source_attribution": 1' in result.stdout


def test_export_canonical_workbook_command_json_reports_errors(tmp_path):
    workbook_path = tmp_path / "canonical.xlsx"
    workbook = Workbook()
    species = workbook.active
    species.title = "Species_Master"
    species.append(["Species_ID", "Botanical_Name", "Evidence_Confidence"])
    species.append(["BAD-0001", "Achillea millefolium", "A"])
    workbook.save(workbook_path)

    result = runner.invoke(
        app,
        [
            "export-canonical-workbook",
            str(workbook_path),
            "--out-dir",
            str(tmp_path / "out"),
            "--json",
        ],
    )

    assert result.exit_code == 1
    assert "invalid_species_id" in result.stdout
    assert (tmp_path / "out" / "species.csv").exists()


def test_validate_score_inputs_command_json_reports_errors(tmp_path):
    scores_path = tmp_path / "scores.csv"
    scores_path.write_text(
        "species_id,score_family,metric,value,source_id,evidence_confidence,review_status\n"
        "BCNPPD-0001,PSI,Native Bee Score,4,SRC-0001,A,pending_review\n",
        encoding="utf-8",
    )

    result = runner.invoke(app, ["validate-score-inputs", str(scores_path), "--json"])

    assert result.exit_code == 1
    assert "unreviewed_score_input" in result.stdout
    assert "missing_score_weight" in result.stdout


def test_calculate_scores_command_json(tmp_path):
    scores_path = tmp_path / "scores.json"
    scores_path.write_text(
        json.dumps(
            [
                {
                    "species_id": "BCNPPD-0001",
                    "score_family": "UNI",
                    "metric": "Urban Toughness",
                    "value": "4",
                    "source_id": "SRC-0001",
                    "evidence_confidence": "A",
                    "review_status": "accepted",
                }
            ]
        ),
        encoding="utf-8",
    )
    weights_path = tmp_path / "weights.json"
    weights_path.write_text(
        json.dumps(
            [
                {
                    "score_family": "UNI",
                    "metric": "Urban Toughness",
                    "weight": "1",
                }
            ]
        ),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "calculate-scores",
            str(scores_path),
            "--weights",
            str(weights_path),
            "--json",
        ],
    )

    assert result.exit_code == 0
    assert '"score": 80.0' in result.stdout
    assert '"score_family": "UNI"' in result.stdout


def test_validate_foundation_command_json():
    result = runner.invoke(
        app,
        ["validate-foundation", "data/foundation/v1.0.0a", "--json"],
    )

    assert result.exit_code == 0
    assert '"species": 1' in result.stdout
    assert '"diagnostics": []' in result.stdout


def test_generate_vancouver_poc_list_command_json(tmp_path):
    result = runner.invoke(
        app,
        [
            "generate-vancouver-poc-list",
            "data/workbooks/native_plant_restoration_workbook_v1.0.0c.xlsx",
            "--out-dir",
            str(tmp_path / "poc"),
            "--json",
        ],
    )

    assert result.exit_code == 0
    assert '"plant_list": 20' in result.stdout
    assert (tmp_path / "poc" / "plant_list.csv").exists()


def test_validate_vancouver_poc_list_command_json():
    result = runner.invoke(
        app,
        ["validate-vancouver-poc-list", "data/poc/vancouver", "--json"],
    )

    assert result.exit_code == 0
    assert '"source_attribution": 41' in result.stdout
    assert '"diagnostics": []' in result.stdout


def test_harden_vancouver_evidence_command_json(tmp_path):
    result = runner.invoke(
        app,
        [
            "harden-vancouver-evidence",
            "data/poc/vancouver",
            "--out-dir",
            str(tmp_path / "hardening"),
            "--json",
        ],
    )

    assert result.exit_code == 0
    assert '"reviewed_fields": 80' in result.stdout
    assert '"score_readiness": 60' in result.stdout
    assert (tmp_path / "hardening" / "score_readiness.csv").exists()


def test_validate_vancouver_evidence_command_json():
    result = runner.invoke(
        app,
        ["validate-vancouver-evidence", "data/poc/vancouver/evidence_hardening", "--json"],
    )

    assert result.exit_code == 0
    assert '"hardened_plant_list": 20' in result.stdout
    assert '"diagnostics": []' in result.stdout


def test_generate_vancouver_usability_command_json(tmp_path):
    result = runner.invoke(
        app,
        [
            "generate-vancouver-usability",
            "data/poc/vancouver/evidence_hardening",
            "--out-dir",
            str(tmp_path / "usability"),
            "--json",
        ],
    )

    assert result.exit_code == 0
    assert '"plant_table": 20' in result.stdout
    assert '"view_summary": 6' in result.stdout
    assert (tmp_path / "usability" / "index.html").exists()


def test_validate_vancouver_usability_command_json():
    result = runner.invoke(
        app,
        ["validate-vancouver-usability", "data/poc/vancouver/usability", "--json"],
    )

    assert result.exit_code == 0
    assert '"use_case_views": 61' in result.stdout
    assert '"diagnostics": []' in result.stdout
