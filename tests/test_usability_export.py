from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from typer.testing import CliRunner

from bc_npp_database.cli import app
from bc_npp_database.usability_export import (
    export_vancouver_poc_excel,
    has_error_diagnostics,
    validate_vancouver_poc_excel,
)

POC_DIR = Path("data/poc/vancouver")
runner = CliRunner()


def test_export_vancouver_poc_excel_creates_formatted_workbook(tmp_path):
    out_path = tmp_path / "vancouver_poc.xlsx"

    result = export_vancouver_poc_excel(POC_DIR, out_path)

    assert not has_error_diagnostics(result.diagnostics)
    assert result.counts["plant_list"] == 493
    assert result.counts["source_attribution"] == 5176
    assert result.counts["provider_approvals"] == 5160
    assert result.counts["provider_candidates"] == 402
    assert result.counts["provider_attributes"] == 1816
    assert result.counts["pollinator_gaps"] == 3451
    assert out_path.exists()

    workbook = load_workbook(out_path)
    try:
        assert workbook.sheetnames[:4] == ["Overview", "Manifests", "Plant List", "Sources"]
        assert "Provider Approvals" in workbook.sheetnames
        assert "Pollinator Review" in workbook.sheetnames

        overview = workbook["Overview"]
        assert overview["A1"].value == "BC-NPPD Vancouver PoC Export"
        assert "candidate/review data" in overview["B5"].value
        assert overview["A8"].hyperlink.target == "#'Plant List'!A1"

        plant_list = workbook["Plant List"]
        assert plant_list.freeze_panes == "A3"
        assert plant_list.auto_filter.ref is None
        assert plant_list["A2"].fill.fgColor.rgb == "001F4E79"
        assert plant_list["A2"].font.bold is True
        assert len(plant_list.tables) == 1
        assert next(iter(plant_list.tables.values())).ref == "A2:P495"

        sources = workbook["Sources"]
        assert sources["E3"].hyperlink.target.startswith("https://")
    finally:
        workbook.close()

    _assert_no_standalone_sheet_autofilters(out_path)


def test_validate_vancouver_poc_excel_reports_missing_workbook(tmp_path):
    result = validate_vancouver_poc_excel(tmp_path / "missing.xlsx")

    assert {diagnostic.code for diagnostic in result.diagnostics} == {
        "missing_vancouver_poc_excel"
    }


def test_export_vancouver_poc_excel_command_json(tmp_path):
    out_path = tmp_path / "vancouver_poc.xlsx"

    result = runner.invoke(
        app,
        [
            "export-vancouver-poc-excel",
            str(POC_DIR),
            "--out-path",
            str(out_path),
            "--json",
        ],
    )

    assert result.exit_code == 0
    assert '"plant_list": 493' in result.stdout
    assert '"workbook":' in result.stdout
    assert out_path.exists()

    validate_result = runner.invoke(
        app,
        ["validate-vancouver-poc-excel", str(out_path), "--json"],
    )

    assert validate_result.exit_code == 0
    assert '"worksheets":' in validate_result.stdout


def _assert_no_standalone_sheet_autofilters(path: Path) -> None:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        worksheet_names = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ]
        for worksheet_name in worksheet_names:
            root = ElementTree.fromstring(archive.read(worksheet_name))
            assert root.find("main:autoFilter", namespace) is None
