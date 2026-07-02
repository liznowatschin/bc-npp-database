import csv
from importlib import resources

from openpyxl import Workbook

from bc_npp_database.canonical import (
    CanonicalSpeciesRecord,
    export_canonical_tables,
    import_canonical_workbook,
    load_lookup_seed,
    master_species_columns,
    normalize_header,
)


def _write_canonical_workbook(path, *, invalid=False, excluded_url=""):
    workbook = Workbook()
    species = workbook.active
    species.title = "Species_Master"
    species.append(
        [
            "Species_ID",
            "Botanical_Name",
            "Common_Name",
            "Evidence_Confidence",
            "Primary_References",
            "Notes",
        ]
    )
    species.append(
        [
            "CDF-0001" if invalid else "BCNPPD-0001",
            "" if invalid else "Achillea millefolium",
            "yarrow",
            "Certain" if invalid else "A",
            "SRC-0001",
            excluded_url,
        ]
    )

    lookups = workbook.create_sheet("Lookup_Tables")
    lookups.append(["lookup_name", "value"])
    lookups.append(["Life Cycle", "Perennial"])

    attribution = workbook.create_sheet("Source_Attribution")
    attribution.append(["Source ID", "Species_ID", "Field", "Confidence", "Notes"])
    attribution.append(["SRC-0001", "BCNPPD-0001", "habitat", "B", excluded_url])

    bloom = workbook.create_sheet("Bloom_Calendar")
    bloom.append(["Species_ID", "Common_Name", "Jan", "Feb", "Mar"])
    bloom.append(["BCNPPD-0001", "yarrow", "", "", "x"])

    workbook.create_sheet("Dashboard")
    workbook.save(path)
    return path


def test_schema_loading_preserves_master_column_order():
    columns = master_species_columns()

    assert columns[:4] == ("Species ID", "Botanical Name", "Accepted Name", "Synonyms")
    assert "Evidence Level" in columns
    assert load_lookup_seed()[0].to_dict() == {"lookup_name": "Life Cycle", "value": "Annual"}


def test_packaged_schema_resources_are_available():
    schema = resources.files("bc_npp_database").joinpath(
        "schemas",
        "master_species_columns.csv",
    )

    assert schema.is_file()


def test_header_aliases_and_species_serialization():
    columns = ("Species ID", "Botanical Name", "Evidence Level")
    record = CanonicalSpeciesRecord.from_mapping(
        {
            "Species_ID": "BCNPPD-0001",
            "Botanical_Name": "Achillea millefolium",
            "Evidence_Confidence": "A",
        },
        columns,
    )

    assert normalize_header("Species_ID") == "Species ID"
    assert record.species_id == "BCNPPD-0001"
    assert record.to_dict(columns) == {
        "Species ID": "BCNPPD-0001",
        "Botanical Name": "Achillea millefolium",
        "Evidence Level": "A",
    }


def test_import_canonical_workbook_reads_approved_sheets(tmp_path):
    path = _write_canonical_workbook(tmp_path / "fixture.xlsx")

    result = import_canonical_workbook(path)

    assert len(result.species) == 1
    assert len(result.lookups) == 1
    assert len(result.source_attribution) == 1
    assert len(result.bloom_calendar) == 1
    assert {diagnostic.code for diagnostic in result.diagnostics} == {"unsupported_sheet"}


def test_import_canonical_workbook_reports_invalid_species_fields(tmp_path):
    path = _write_canonical_workbook(tmp_path / "fixture.xlsx", invalid=True)

    result = import_canonical_workbook(path)

    codes = {diagnostic.code for diagnostic in result.diagnostics}
    assert {"invalid_species_id", "invalid_evidence_confidence", "missing_required_field"} <= codes


def test_import_canonical_workbook_detects_excluded_sources(tmp_path):
    path = _write_canonical_workbook(
        tmp_path / "fixture.xlsx",
        excluded_url="https://vancouver.ca/files/cov/vancouver-gri-planting-guidelines.pdf",
    )

    result = import_canonical_workbook(path)

    assert any(diagnostic.code == "excluded_source" for diagnostic in result.diagnostics)


def test_export_canonical_tables_is_deterministic(tmp_path):
    workbook_path = _write_canonical_workbook(tmp_path / "fixture.xlsx")
    result = import_canonical_workbook(workbook_path)

    export = export_canonical_tables(result, tmp_path / "out")

    assert sorted(export.paths) == [
        "bloom_calendar",
        "diagnostics",
        "lookups",
        "source_attribution",
        "species",
    ]
    with open(export.paths["species"], newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        row = next(reader)
    assert header[:2] == ["Species ID", "Botanical Name"]
    assert row[:2] == ["BCNPPD-0001", "Achillea millefolium"]
