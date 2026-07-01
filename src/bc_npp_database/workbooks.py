"""Workbook inspection helpers for BC-NPPD."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .diagnostics import Diagnostic, Severity


@dataclass(frozen=True)
class SheetInventory:
    """Inventory details for a workbook sheet."""

    name: str
    max_row: int | None
    max_column: int | None
    headers: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable sheet inventory dictionary."""
        return {
            "name": self.name,
            "max_row": self.max_row,
            "max_column": self.max_column,
            "headers": list(self.headers),
        }


@dataclass(frozen=True)
class WorkbookInventory:
    """Read-only workbook inventory."""

    path: str
    sheets: tuple[SheetInventory, ...]

    def to_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable workbook inventory dictionary."""
        return {
            "path": self.path,
            "sheets": [sheet.to_dict() for sheet in self.sheets],
        }


EXPECTED_WORKBOOK_SHEETS = (
    "Species_Master",
    "Lookup_Tables",
    "Reference_Policy",
    "Source_Attribution",
    "Bloom_Calendar",
    "QA_Log",
)

CANONICAL_TABLES_BY_SHEET = {
    "Species_Master": "candidate_species",
    "Lookup_Tables": "controlled_vocabularies",
    "Reference_Policy": "source_policy",
    "Source_Attribution": "source_attribution",
    "Bloom_Calendar": "bloom_calendar",
    "Dashboard": "report_output",
    "QA_Log": "review_queue",
}


def inventory_workbook(path: Path) -> WorkbookInventory:
    """Return a read-only workbook inventory."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[SheetInventory] = []
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            first_row = next(rows, ())
            headers = tuple("" if value is None else str(value).strip() for value in first_row)
            sheets.append(
                SheetInventory(
                    name=worksheet.title,
                    max_row=worksheet.max_row,
                    max_column=worksheet.max_column,
                    headers=headers,
                )
            )
        return WorkbookInventory(path=str(path), sheets=tuple(sheets))
    finally:
        workbook.close()


def validate_workbook(path: Path) -> list[Diagnostic]:
    """Validate workbook readability and expected foundation sheets."""
    try:
        inventory = inventory_workbook(path)
    except (InvalidFileException, OSError, KeyError, ValueError) as exc:
        return [
            Diagnostic(
                code="workbook_unreadable",
                message=f"Workbook could not be read: {exc}",
                severity=Severity.ERROR,
                context={"path": str(path)},
            )
        ]

    diagnostics: list[Diagnostic] = []
    sheet_names = {sheet.name for sheet in inventory.sheets}
    for expected in EXPECTED_WORKBOOK_SHEETS:
        if expected not in sheet_names:
            diagnostics.append(
                Diagnostic(
                    code="missing_sheet",
                    message=f"Expected workbook sheet is missing: {expected}",
                    severity=Severity.ERROR,
                    field="sheet",
                    value=expected,
                    context={"path": str(path)},
                )
            )

    for sheet in inventory.sheets:
        if sheet.name in {"Species_Master", "Source_Attribution"} and not any(sheet.headers):
            diagnostics.append(
                Diagnostic(
                    code="missing_header",
                    message=f"Sheet has no header row: {sheet.name}",
                    severity=Severity.ERROR,
                    field="sheet",
                    value=sheet.name,
                    context={"path": str(path)},
                )
            )
    return diagnostics
