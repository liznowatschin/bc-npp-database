"""Excel workbook export for Vancouver PoC artifacts."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .diagnostics import Diagnostic, Severity
from .vancouver_poc import (
    has_error_diagnostics as has_vancouver_poc_error_diagnostics,
)
from .vancouver_poc import (
    validate_vancouver_poc_list,
)

EXPORT_CAVEAT = (
    "Excel export is a formatted inspection copy of Vancouver PoC artifacts. "
    "Provider, usability, evidence, and pollinator rows remain candidate/review "
    "data unless their source artifacts explicitly say otherwise."
)

SHEET_SPECS = (
    ("Plant List", "plant_list.csv", "Core Vancouver PoC plant records."),
    ("Sources", "sources.csv", "Source registry."),
    ("Source Attribution", "source_attribution.csv", "Claim-to-source links."),
    ("Requested Additions", "requested_species_additions.csv", "Submitted species audit."),
    ("Diagnostics", "diagnostics.csv", "Core PoC diagnostics and caveats."),
    ("Provider Approvals", "provider_data/approval_manifest.csv", "Provider review decisions."),
    ("Provider Candidates", "provider_data/candidate_species.csv", "Provider candidate species."),
    (
        "Provider Attributes",
        "provider_data/candidate_attributes.csv",
        "Provider candidate attributes.",
    ),
    ("Provider Suppliers", "provider_data/supplier_availability.csv", "Supplier availability."),
    ("Provider Mowability", "provider_data/mowability.csv", "Provisional mowability observations."),
    ("Provider Attribution", "provider_data/source_attribution.csv", "Provider attribution rows."),
    ("Evidence Plants", "evidence_hardening/hardened_plant_list.csv", "Evidence-hardened plants."),
    ("Evidence Gaps", "evidence_hardening/evidence_gap_report.csv", "Field-level evidence gaps."),
    ("Score Readiness", "evidence_hardening/score_readiness.csv", "Score readiness checks."),
    ("Reviewed Sources", "evidence_hardening/reviewed_sources.csv", "Reviewed source registry."),
    ("Reviewed Fields", "evidence_hardening/reviewed_fields.csv", "Reviewed field flags."),
    ("Usability Plants", "usability/plant_table.csv", "P8 usability table."),
    ("Use Case Views", "usability/use_case_views.csv", "P8 use-case membership rows."),
    ("View Summary", "usability/view_summary.csv", "P8 use-case summary rows."),
    ("Pollinator Review", "pollinator_module/pollinator_review.csv", "Pollinator review queue."),
    (
        "Pollinator Gaps",
        "pollinator_module/pollinator_evidence_gaps.csv",
        "Pollinator evidence gaps.",
    ),
    (
        "Pollinator Sources",
        "pollinator_module/pollinator_source_requirements.csv",
        "Pollinator source requirements.",
    ),
)

MANIFEST_SPECS = (
    ("Core Manifest", "manifest.json", "Core Vancouver PoC manifest."),
    ("Provider Manifest", "provider_data/manifest.json", "Provider data manifest."),
    ("Evidence Manifest", "evidence_hardening/manifest.json", "Evidence-hardening manifest."),
    ("Usability Manifest", "usability/manifest.json", "Usability manifest."),
    ("Pollinator Manifest", "pollinator_module/manifest.json", "Pollinator module manifest."),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN_BORDER = Border(bottom=Side(style="thin", color="B7B7B7"))
HYPERLINK_FONT = Font(color="0563C1", underline="single")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class VancouverPocExcelResult:
    """Vancouver PoC Excel export summary."""

    path: str
    counts: dict[str, int]
    diagnostics: tuple[Diagnostic, ...]
    paths: dict[str, str]

    def to_summary_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable summary."""
        return {
            "path": self.path,
            "counts": self.counts,
            "paths": self.paths,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


def export_vancouver_poc_excel(
    poc_dir: Path,
    output_path: Path,
    *,
    title: str = "BC-NPPD Vancouver PoC Export",
) -> VancouverPocExcelResult:
    """Export Vancouver PoC artifacts to a formatted Excel workbook."""
    poc_dir = poc_dir.resolve()
    output_path = output_path.resolve()
    diagnostics: list[Diagnostic] = []

    validation = validate_vancouver_poc_list(poc_dir)
    diagnostics.extend(validation.diagnostics)
    if has_vancouver_poc_error_diagnostics(validation.diagnostics):
        return VancouverPocExcelResult(
            path=str(output_path),
            counts=_empty_counts(),
            diagnostics=tuple(diagnostics),
            paths={},
        )

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"

    sheet_counts: dict[str, int] = {}
    sheet_paths: dict[str, str] = {}

    manifest_rows = _manifest_rows(poc_dir)
    _write_overview_sheet(overview, title, poc_dir, output_path)
    _write_manifest_sheet(workbook, manifest_rows)

    for sheet_name, relative_path, description in SHEET_SPECS:
        source_path = poc_dir / relative_path
        if not source_path.exists():
            continue
        rows, fieldnames = _read_csv(source_path, diagnostics)
        sheet = workbook.create_sheet(sheet_name)
        _write_data_sheet(sheet, fieldnames, rows, description)
        sheet_counts[_count_key(sheet_name)] = len(rows)
        sheet_paths[_count_key(sheet_name)] = str(source_path)

    _write_counts_to_overview(overview, sheet_counts, sheet_paths)
    _style_workbook(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    diagnostics.append(
        Diagnostic(
            code="vancouver_poc_excel_exported",
            message=EXPORT_CAVEAT,
            severity=Severity.WARNING,
            context={"sheet_count": len(workbook.sheetnames), "output_path": str(output_path)},
        )
    )
    counts = {
        "worksheets": len(workbook.sheetnames),
        "data_sheets": len(sheet_counts),
        **sheet_counts,
    }
    paths = {"workbook": str(output_path), **sheet_paths}
    return VancouverPocExcelResult(str(output_path), counts, tuple(diagnostics), paths)


def validate_vancouver_poc_excel(path: Path) -> VancouverPocExcelResult:
    """Validate that a Vancouver PoC Excel export is readable and structured."""
    diagnostics: list[Diagnostic] = []
    counts = _empty_counts()
    if not path.exists():
        diagnostics.append(
            _diagnostic("missing_vancouver_poc_excel", field="path", value=str(path))
        )
        return VancouverPocExcelResult(str(path), counts, tuple(diagnostics), {})
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except OSError as exc:
        diagnostics.append(
            _diagnostic("vancouver_poc_excel_unreadable", field="path", value=f"{path}: {exc}")
        )
        return VancouverPocExcelResult(str(path), counts, tuple(diagnostics), {})
    try:
        sheetnames = workbook.sheetnames
        counts["worksheets"] = len(sheetnames)
        counts["data_sheets"] = len(
            [name for name in sheetnames if name not in {"Overview", "Manifests"}]
        )
        for required in ("Overview", "Manifests", "Plant List", "Sources", "Source Attribution"):
            if required not in sheetnames:
                diagnostics.append(_diagnostic("missing_vancouver_poc_excel_sheet", value=required))
        for worksheet in workbook.worksheets:
            if worksheet.title == "Overview":
                continue
            counts[_count_key(worksheet.title)] = max(worksheet.max_row - 2, 0)
    finally:
        workbook.close()
    return VancouverPocExcelResult(str(path), counts, tuple(diagnostics), {"workbook": str(path)})


def has_error_diagnostics(diagnostics: tuple[Diagnostic, ...]) -> bool:
    """Return whether diagnostics contain at least one error."""
    return any(diagnostic.severity == Severity.ERROR for diagnostic in diagnostics)


def _write_overview_sheet(
    worksheet,
    title: str,
    poc_dir: Path,
    output_path: Path,
) -> None:
    worksheet.append([title])
    worksheet.append(["Generated UTC", datetime.now(UTC).isoformat(timespec="seconds")])
    worksheet.append(["Input PoC directory", str(poc_dir)])
    worksheet.append(["Output workbook", str(output_path)])
    worksheet.append(["Caveat", EXPORT_CAVEAT])
    worksheet.append([])
    worksheet.append(["Included Sheet", "Rows", "Source CSV"])


def _write_counts_to_overview(
    worksheet,
    sheet_counts: dict[str, int],
    sheet_paths: dict[str, str],
) -> None:
    start_row = 8
    for offset, (key, count) in enumerate(sheet_counts.items()):
        sheet_name = _sheet_label_from_key(key)
        row_number = start_row + offset
        worksheet.cell(row=row_number, column=1, value=sheet_name)
        worksheet.cell(row=row_number, column=2, value=count)
        worksheet.cell(row=row_number, column=3, value=sheet_paths[key])
        worksheet.cell(row=row_number, column=1).hyperlink = f"#'{sheet_name}'!A1"
        worksheet.cell(row=row_number, column=1).style = "Hyperlink"


def _write_manifest_sheet(workbook: Workbook, manifest_rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet("Manifests")
    _write_data_sheet(
        worksheet,
        ["manifest", "field", "value"],
        manifest_rows,
        "Flattened manifest metadata from Vancouver PoC artifact directories.",
    )


def _write_data_sheet(
    worksheet,
    fieldnames: list[str],
    rows: list[dict[str, str]],
    description: str,
) -> None:
    worksheet.append([description])
    worksheet.append(fieldnames)
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])
    worksheet.freeze_panes = "A3"
    _add_table_if_possible(worksheet, len(fieldnames), len(rows))
    _add_hyperlinks(worksheet, fieldnames)


def _style_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = TITLE_FILL
                cell.font = Font(bold=True, color="274E13")
                cell.alignment = WRAP_ALIGNMENT
        header_row = 2 if worksheet.title != "Overview" else 7
        if worksheet.max_row >= header_row:
            for cell in worksheet[header_row]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if worksheet.title == "Overview":
            for row in worksheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2):
                for cell in row:
                    cell.alignment = WRAP_ALIGNMENT
            for cell in worksheet[9]:
                cell.fill = SUBHEADER_FILL
                cell.font = Font(bold=True)
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.row > header_row:
                    cell.alignment = WRAP_ALIGNMENT
        _size_columns(worksheet)


def _add_table_if_possible(worksheet, column_count: int, row_count: int) -> None:
    if column_count == 0 or row_count == 0:
        return
    end_column = get_column_letter(column_count)
    table_ref = f"A2:{end_column}{row_count + 2}"
    display_name = "tbl_" + "".join(ch for ch in worksheet.title if ch.isalnum())
    table = Table(displayName=display_name[:250], ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _add_hyperlinks(worksheet, fieldnames: list[str]) -> None:
    url_fields = {
        index + 1
        for index, field in enumerate(fieldnames)
        if "url" in field.casefold() or field.casefold() in {"citation"}
    }
    for column_index in url_fields:
        for row in range(3, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=column_index)
            value = str(cell.value or "")
            if value.startswith(("http://", "https://")):
                cell.hyperlink = value
                cell.font = HYPERLINK_FONT


def _size_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        values = [str(cell.value or "") for cell in column_cells[:100]]
        max_length = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 48)


def _manifest_rows(poc_dir: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for label, relative_path, description in MANIFEST_SPECS:
        path = poc_dir / relative_path
        if not path.exists():
            continue
        payload = _load_json(path)
        rows.append({"manifest": label, "field": "_path", "value": str(path)})
        rows.append({"manifest": label, "field": "_description", "value": description})
        for field, value in _flatten_json(payload):
            rows.append({"manifest": label, "field": field, "value": value})
    return rows


def _flatten_json(payload: dict[str, Any], prefix: str = "") -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for key, value in payload.items():
        field = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(_flatten_json(value, field))
        elif isinstance(value, list):
            rows.append((field, json.dumps(value, sort_keys=True)))
        else:
            rows.append((field, "" if value is None else str(value)))
    return rows


def _read_csv(path: Path, diagnostics: list[Diagnostic]) -> tuple[list[dict[str, str]], list[str]]:
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            rows = [dict(row) for row in reader]
            return rows, list(reader.fieldnames or [])
    except OSError as exc:
        diagnostics.append(
            _diagnostic("vancouver_poc_excel_csv_unreadable", value=f"{path}: {exc}")
        )
        return [], []


def _load_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _empty_counts() -> dict[str, int]:
    return {"worksheets": 0, "data_sheets": 0}


def _count_key(sheet_name: str) -> str:
    return sheet_name.casefold().replace(" ", "_").replace("-", "_")


def _sheet_label_from_key(key: str) -> str:
    for sheet_name, _, _ in SHEET_SPECS:
        if _count_key(sheet_name) == key:
            return sheet_name
    return key.replace("_", " ").title()


def _diagnostic(
    code: str,
    *,
    field: str | None = None,
    value: str | None = None,
) -> Diagnostic:
    return Diagnostic(
        code=code,
        message=code.replace("_", " ").capitalize() + ".",
        severity=Severity.ERROR,
        field=field,
        value=value,
    )
