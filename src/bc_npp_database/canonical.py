"""Canonical table import and export helpers for BC-NPPD."""

from __future__ import annotations

import csv
import re
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from importlib import resources
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .config import EVIDENCE_CONFIDENCE_VALUES, EXCLUDED_SOURCE_URLS
from .diagnostics import Diagnostic, Severity
from .sources import SPECIES_ID_PATTERN, validate_source_attribution_records
from .validate import diagnose_duplicate_species_ids

APPROVED_IMPORT_SHEETS = {
    "Species_Master",
    "Lookup_Tables",
    "Source_Attribution",
    "Bloom_Calendar",
}

MONTH_COLUMNS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")

SPECIES_REQUIRED_COLUMNS = ("Species ID", "Botanical Name")

SPECIES_HEADER_ALIASES = {
    "Species_ID": "Species ID",
    "Botanical_Name": "Botanical Name",
    "Accepted_Name": "Accepted Name",
    "Common_Name": "Common Name",
    "Native_Status": "Native Status",
    "Plant_Type": "Plant Type",
    "Life_Cycle": "Life Cycle",
    "Growth_Form": "Growth Habit",
    "Bloom_Start": "Bloom Start",
    "Bloom_End": "Bloom End",
    "Flower_Color": "Flower Colour",
    "Light": "Sun",
    "Moisture": "Soil Moisture",
    "Soil": "Soil Texture",
    "Nectar_Value": "Native Bee Score",
    "Bumble_Bee_Value": "Bumble Bee Score",
    "Solitary_Bee_Value": "Mason Bee Score",
    "Butterfly_Value": "Butterfly Nectar Score",
    "Hoverfly_Value": "Hoverfly Score",
    "Seeds_per_kg": "Seeds per kg",
    "Stratification": "Cold Stratification",
    "Urban_Suitability": "Urban Toughness",
    "Evidence_Confidence": "Evidence Level",
    "Primary_References": "Primary References",
}

ATTRIBUTION_HEADER_ALIASES = {
    "Species_ID": "species_id",
    "Species ID": "species_id",
    "Source ID": "source_id",
    "Reference ID": "source_id",
    "Field": "claim_field",
    "Claim Field": "claim_field",
    "Value": "claim_value",
    "Claim Value": "claim_value",
    "Use": "claim_scope",
    "Confidence": "evidence_confidence",
    "Evidence Confidence": "evidence_confidence",
    "Evidence Level": "evidence_confidence",
    "External ID": "external_id",
    "Review Status": "review_status",
    "Notes": "notes",
}


@dataclass(frozen=True)
class SchemaColumn:
    """One canonical schema column definition."""

    section: str
    column_name: str
    data_type: str
    allowed_values_or_notes: str

    @classmethod
    def from_mapping(cls, row: Mapping[str, object]) -> SchemaColumn:
        """Build a schema column from a CSV mapping."""
        return cls(
            section=_clean(row.get("section")),
            column_name=_clean(row.get("column_name")),
            data_type=_clean(row.get("data_type")),
            allowed_values_or_notes=_clean(row.get("allowed_values_or_notes")),
        )

    def to_dict(self) -> dict[str, str]:
        """Return a JSON-serializable schema column dictionary."""
        return {
            "section": self.section,
            "column_name": self.column_name,
            "data_type": self.data_type,
            "allowed_values_or_notes": self.allowed_values_or_notes,
        }


@dataclass(frozen=True)
class CanonicalSpeciesRecord:
    """One canonical candidate species row."""

    values: dict[str, str]

    @property
    def species_id(self) -> str:
        """Return the stable BC-NPPD species ID."""
        return self.values.get("Species ID", "")

    @property
    def botanical_name(self) -> str:
        """Return the botanical name."""
        return self.values.get("Botanical Name", "")

    @classmethod
    def from_mapping(
        cls,
        row: Mapping[str, object],
        columns: Sequence[str],
    ) -> CanonicalSpeciesRecord:
        """Build a species record from workbook or CSV-shaped fields."""
        normalized = normalize_species_row(row)
        return cls({column: normalized.get(column, "") for column in columns})

    def to_dict(self, columns: Sequence[str] | None = None) -> dict[str, str]:
        """Return a stable dictionary with optional canonical column order."""
        if columns is None:
            return dict(self.values)
        return {column: self.values.get(column, "") for column in columns}


@dataclass(frozen=True)
class CanonicalLookupRecord:
    """One controlled vocabulary lookup row."""

    lookup_name: str
    value: str

    @classmethod
    def from_mapping(cls, row: Mapping[str, object]) -> CanonicalLookupRecord:
        """Build a lookup record from common lookup field names."""
        return cls(
            lookup_name=_first(row, "lookup_name", "Lookup Name", "lookup", "Lookup"),
            value=_first(row, "value", "Value"),
        )

    def to_dict(self) -> dict[str, str]:
        """Return a JSON-serializable lookup record dictionary."""
        return {"lookup_name": self.lookup_name, "value": self.value}


@dataclass(frozen=True)
class CanonicalBloomRecord:
    """One bloom-calendar row."""

    species_id: str
    common_name: str
    months: dict[str, str]

    @classmethod
    def from_mapping(cls, row: Mapping[str, object]) -> CanonicalBloomRecord:
        """Build a bloom record from workbook fields."""
        return cls(
            species_id=_first(row, "Species ID", "Species_ID", "species_id"),
            common_name=_first(row, "Common Name", "Common_Name", "common_name"),
            months={month: _first(row, month) for month in MONTH_COLUMNS},
        )

    def to_dict(self) -> dict[str, str]:
        """Return a stable bloom-calendar dictionary."""
        return {
            "Species ID": self.species_id,
            "Common Name": self.common_name,
            **{month: self.months.get(month, "") for month in MONTH_COLUMNS},
        }


@dataclass(frozen=True)
class CanonicalImportResult:
    """Canonical workbook import result."""

    path: str
    species: tuple[CanonicalSpeciesRecord, ...]
    lookups: tuple[CanonicalLookupRecord, ...]
    source_attribution: tuple[dict[str, str], ...]
    bloom_calendar: tuple[CanonicalBloomRecord, ...]
    diagnostics: tuple[Diagnostic, ...]

    def to_summary_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable import summary."""
        return {
            "path": self.path,
            "counts": {
                "species": len(self.species),
                "lookups": len(self.lookups),
                "source_attribution": len(self.source_attribution),
                "bloom_calendar": len(self.bloom_calendar),
                "diagnostics": len(self.diagnostics),
            },
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


@dataclass(frozen=True)
class CanonicalExportResult:
    """Canonical CSV export result."""

    output_dir: str
    paths: dict[str, str]
    diagnostics: tuple[Diagnostic, ...]

    def to_summary_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable export summary."""
        return {
            "output_dir": self.output_dir,
            "paths": self.paths,
            "diagnostics": [diagnostic.to_dict() for diagnostic in self.diagnostics],
        }


def load_master_species_schema(path: Path | None = None) -> tuple[SchemaColumn, ...]:
    """Load canonical master species schema columns in declared order."""
    schema_path = path or _schema_path("master_species_columns.csv")
    with schema_path.open(newline="", encoding="utf-8-sig") as handle:
        return tuple(SchemaColumn.from_mapping(row) for row in csv.DictReader(handle))


def load_lookup_seed(path: Path | None = None) -> tuple[CanonicalLookupRecord, ...]:
    """Load lookup seed rows in declared order."""
    lookup_path = path or _schema_path("lookups_seed.csv")
    with lookup_path.open(newline="", encoding="utf-8-sig") as handle:
        return tuple(CanonicalLookupRecord.from_mapping(row) for row in csv.DictReader(handle))


def master_species_columns(path: Path | None = None) -> tuple[str, ...]:
    """Return canonical master species column names in stable order."""
    return tuple(column.column_name for column in load_master_species_schema(path))


def normalize_header(value: object) -> str:
    """Normalize a workbook header into a canonical-ish field name."""
    header = _clean(value)
    if header in SPECIES_HEADER_ALIASES:
        return SPECIES_HEADER_ALIASES[header]
    compacted = re.sub(r"\s+", " ", header.replace("_", " ")).strip()
    return SPECIES_HEADER_ALIASES.get(compacted, compacted)


def normalize_species_row(row: Mapping[str, object]) -> dict[str, str]:
    """Normalize species row headers and values."""
    normalized: dict[str, str] = {}
    for key, value in row.items():
        header = normalize_header(key)
        if header:
            normalized[header] = _clean(value)
    return normalized


def normalize_source_attribution_row(row: Mapping[str, object]) -> dict[str, str]:
    """Normalize source-attribution headers for P2 validation."""
    normalized: dict[str, str] = {}
    for key, value in row.items():
        header = _clean(key)
        if not header:
            continue
        normalized[ATTRIBUTION_HEADER_ALIASES.get(header, header)] = _clean(value)
    return normalized


def import_canonical_workbook(path: Path) -> CanonicalImportResult:
    """Import approved workbook sheets into canonical in-memory records."""
    schema_columns = master_species_columns()
    diagnostics: list[Diagnostic] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError, ValueError) as exc:
        return CanonicalImportResult(
            path=str(path),
            species=(),
            lookups=(),
            source_attribution=(),
            bloom_calendar=(),
            diagnostics=(
                Diagnostic(
                    code="workbook_unreadable",
                    message=f"Workbook could not be read: {exc}",
                    severity=Severity.ERROR,
                    context={"path": str(path)},
                ),
            ),
        )

    try:
        species_rows: list[CanonicalSpeciesRecord] = []
        lookup_rows: list[CanonicalLookupRecord] = []
        attribution_rows: list[dict[str, str]] = []
        bloom_rows: list[CanonicalBloomRecord] = []

        for sheet_name in workbook.sheetnames:
            if sheet_name not in APPROVED_IMPORT_SHEETS:
                diagnostics.append(
                    Diagnostic(
                        code="unsupported_sheet",
                        message=f"Sheet is not imported by the canonical pipeline: {sheet_name}",
                        severity=Severity.WARNING,
                        field="sheet",
                        value=sheet_name,
                    )
                )
                continue
            worksheet = workbook[sheet_name]
            rows = list(_worksheet_mappings(worksheet))
            if sheet_name == "Species_Master":
                species_rows = [
                    CanonicalSpeciesRecord.from_mapping(row, schema_columns)
                    for row in rows
                    if _row_has_values(row)
                ]
                diagnostics.extend(_validate_species_records(species_rows, schema_columns))
            elif sheet_name == "Lookup_Tables":
                lookup_rows = _import_lookup_rows(rows)
            elif sheet_name == "Source_Attribution":
                attribution_rows = [
                    normalize_source_attribution_row(row) for row in rows if _row_has_values(row)
                ]
                diagnostics.extend(validate_source_attribution_records(attribution_rows))
            elif sheet_name == "Bloom_Calendar":
                bloom_rows = [
                    CanonicalBloomRecord.from_mapping(row) for row in rows if _row_has_values(row)
                ]
                diagnostics.extend(_validate_bloom_records(bloom_rows))

        return CanonicalImportResult(
            path=str(path),
            species=tuple(species_rows),
            lookups=tuple(lookup_rows),
            source_attribution=tuple(attribution_rows),
            bloom_calendar=tuple(bloom_rows),
            diagnostics=tuple(diagnostics),
        )
    finally:
        workbook.close()


def export_canonical_tables(
    result: CanonicalImportResult,
    output_dir: Path,
) -> CanonicalExportResult:
    """Export canonical import results as deterministic CSV tables."""
    output_dir.mkdir(parents=True, exist_ok=True)
    schema_columns = master_species_columns()
    paths: dict[str, str] = {}

    paths["species"] = str(output_dir / "species.csv")
    _write_csv(
        Path(paths["species"]),
        [record.to_dict(schema_columns) for record in result.species],
        schema_columns,
    )

    paths["lookups"] = str(output_dir / "lookups.csv")
    _write_csv(
        Path(paths["lookups"]),
        [record.to_dict() for record in result.lookups],
        ("lookup_name", "value"),
    )

    attribution_fields = (
        "source_id",
        "species_id",
        "claim_field",
        "claim_value",
        "claim_scope",
        "evidence_confidence",
        "external_id",
        "review_status",
        "notes",
    )
    paths["source_attribution"] = str(output_dir / "source_attribution.csv")
    _write_csv(Path(paths["source_attribution"]), result.source_attribution, attribution_fields)

    bloom_fields = ("Species ID", "Common Name", *MONTH_COLUMNS)
    paths["bloom_calendar"] = str(output_dir / "bloom_calendar.csv")
    _write_csv(
        Path(paths["bloom_calendar"]),
        [record.to_dict() for record in result.bloom_calendar],
        bloom_fields,
    )

    if result.diagnostics:
        paths["diagnostics"] = str(output_dir / "diagnostics.csv")
        _write_csv(
            Path(paths["diagnostics"]),
            [diagnostic.to_dict() for diagnostic in result.diagnostics],
            ("severity", "code", "message", "row_number", "field", "value", "context"),
        )

    return CanonicalExportResult(
        output_dir=str(output_dir),
        paths=paths,
        diagnostics=result.diagnostics,
    )


def has_error_diagnostics(diagnostics: Iterable[Diagnostic]) -> bool:
    """Return whether diagnostics contain at least one error."""
    return any(diagnostic.severity == Severity.ERROR for diagnostic in diagnostics)


def _validate_species_records(
    records: Sequence[CanonicalSpeciesRecord],
    columns: Sequence[str],
) -> list[Diagnostic]:
    diagnostics: list[Diagnostic] = []
    for row_number, record in enumerate(records, start=2):
        values = record.to_dict(columns)
        for column in SPECIES_REQUIRED_COLUMNS:
            if not values.get(column, ""):
                diagnostics.append(
                    Diagnostic(
                        code="missing_required_field",
                        message=f"Missing required canonical species field: {column}",
                        severity=Severity.ERROR,
                        row_number=row_number,
                        field=column,
                    )
                )
        if record.species_id and not SPECIES_ID_PATTERN.fullmatch(record.species_id):
            diagnostics.append(
                Diagnostic(
                    code="invalid_species_id",
                    message="Invalid BC-NPPD species ID.",
                    severity=Severity.ERROR,
                    row_number=row_number,
                    field="Species ID",
                    value=record.species_id,
                )
            )
        evidence = values.get("Evidence Level", "")
        if evidence not in EVIDENCE_CONFIDENCE_VALUES:
            diagnostics.append(
                Diagnostic(
                    code="invalid_evidence_confidence",
                    message="Invalid evidence confidence.",
                    severity=Severity.ERROR,
                    row_number=row_number,
                    field="Evidence Level",
                    value=evidence,
                )
            )
        diagnostics.extend(_diagnose_excluded_values(values, row_number))
    diagnostics.extend(
        diagnose_duplicate_species_ids(record.to_dict(columns) for record in records)
    )
    return diagnostics


def _validate_bloom_records(records: Sequence[CanonicalBloomRecord]) -> list[Diagnostic]:
    diagnostics: list[Diagnostic] = []
    for row_number, record in enumerate(records, start=2):
        if record.species_id and not SPECIES_ID_PATTERN.fullmatch(record.species_id):
            diagnostics.append(
                Diagnostic(
                    code="invalid_species_id",
                    message="Invalid BC-NPPD species ID.",
                    severity=Severity.ERROR,
                    row_number=row_number,
                    field="Species ID",
                    value=record.species_id,
                )
            )
        diagnostics.extend(_diagnose_excluded_values(record.to_dict(), row_number))
    return diagnostics


def _import_lookup_rows(rows: Sequence[Mapping[str, object]]) -> list[CanonicalLookupRecord]:
    imported: list[CanonicalLookupRecord] = []
    for row in rows:
        if {"lookup_name", "value"} <= set(row) or {"Lookup Name", "Value"} <= set(row):
            record = CanonicalLookupRecord.from_mapping(row)
            if record.lookup_name and record.value:
                imported.append(record)
            continue
        for key, value in row.items():
            lookup_name = _clean(key)
            lookup_value = _clean(value)
            if lookup_name and lookup_value and lookup_name.lower() != "definition":
                imported.append(CanonicalLookupRecord(lookup_name=lookup_name, value=lookup_value))
    return imported


def _worksheet_mappings(worksheet: Any) -> list[dict[str, object]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    header_values = [_clean(header) for header in headers]
    mappings: list[dict[str, object]] = []
    for row in rows:
        mapping: dict[str, object] = {}
        for index, header in enumerate(header_values):
            if not header:
                continue
            mapping[header] = row[index] if index < len(row) else None
        mappings.append(mapping)
    return mappings


def _write_csv(path: Path, rows: Iterable[Mapping[str, object]], fieldnames: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _clean(row.get(field)) for field in fieldnames})


def _diagnose_excluded_values(row: Mapping[str, object], row_number: int) -> list[Diagnostic]:
    row_text = " ".join(_clean(value) for value in row.values())
    diagnostics: list[Diagnostic] = []
    for url in sorted(EXCLUDED_SOURCE_URLS):
        if url in row_text:
            diagnostics.append(
                Diagnostic(
                    code="excluded_source",
                    message=f"Excluded source URL found: {url}",
                    severity=Severity.ERROR,
                    row_number=row_number,
                    value=url,
                )
            )
    return diagnostics


def _schema_path(filename: str) -> Path:
    repo_path = Path(__file__).resolve().parents[2] / "schemas" / filename
    if repo_path.exists():
        return repo_path
    resource = resources.files("bc_npp_database").joinpath("schemas", filename)
    with resources.as_file(resource) as path:
        return Path(path)


def _row_has_values(row: Mapping[str, object]) -> bool:
    return any(_clean(value) for value in row.values())


def _first(row: Mapping[str, object], *keys: str) -> str:
    for key in keys:
        if key in row:
            return _clean(row[key])
    return ""


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
